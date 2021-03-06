import arcpy, os,sys
from openpyxl import Workbook
reload(sys)
sys.setdefaultencoding("utf8")
wb = Workbook()

geodatabase=r"E:\ActualizacionSectores\Integrada\DESCONTAMINA_2019.gdb"
salida = r"E:\ActualizacionSectores\Integrada\DESCONTAMINA_2019.xlsx"
arcpy.env.workspace = geodatabase
listadatasets= arcpy.ListDatasets("*")

for Dataset in listadatasets:
    print Dataset
    datasetin= geodatabase+os.sep+Dataset
    arcpy.env.workspace= datasetin
    listfeat= arcpy.ListFeatureClasses("*")
    for fc in listfeat:
        ws = wb.create_sheet(title=fc[:31])
        ws.cell(column=1, row=1, value="{0}".format("Nombre de Campo"))
        ws.cell(column=2, row=1, value="{0}".format("Alias"))
        ws.cell(column=3, row=1, value="{0}".format("Tipo"))
        ws.cell(column=4, row=1, value="{0}".format("Longitud"))
        #ws.cell(column=4, row=1, value="{0}".format("Descripcion"))
        listaCampos=arcpy.ListFields(fc)
        X=1
        for field in listaCampos:

            if field.editable == True and field.type!="Geometry":
                X = X + 1
                ws.cell(column=1, row=X, value="{0}".format(field.name))
                ws.cell(column=2, row=X, value="{0}".format(field.aliasName))
                ws.cell(column=3, row=X, value="{0}".format(field.type))
                ws.cell(column=4, row=X, value="{0}".format(field.length))
                #ws.cell(column=4, row=X, value="{0}".format(field.length))

        print (fc)

wb.save(salida)