import arcpy, os,sys
import arcpy, os, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
wb = Workbook()
dest_filename = r'C:\Temp\Libro1.xlsx'
wb.save(filename = dest_filename)

GeodatabaseTopo=r"C:\Users\APN\Documents\IGAC\INFORMACION PARA FERNANDO\10000\172IIIA1\172IIIA1_Depurada.gdb"
GeodatabaseGeo=r"C:\Users\APN\Documents\IGAC\INFORMACION PARA FERNANDO\10000\172IIIA1_RS_2016_V1.mdb"

def CountMoreZero(Feat):
    X=0
    with arcpy.da.SearchCursor(Feat, "OBJECTID") as cursor:
        for row in cursor:
            if X>0:
                break
            X = X+1
    return X

def recorrerTopo(geodatabase):
    arcpy.env.workspace=geodatabase
    ListaFeat=arcpy.ListFeatureClasses()
    ListaCapasTopo=[]
    for fc in ListaFeat:
        count = CountMoreZero(fc)
        print fc + "..." + str(count)
        if count > 0:
            ListaCapasTopo.append(geodatabase+ os.sep+ fc)
    return ListaCapasTopo

def recorreGeo (geodatabase):
    arcpy.env.workspace = geodatabase
    ListaDatasets= arcpy.ListDatasets("*","Feature")
    ListaCapasGeo =[]
    for dataset in ListaDatasets:
        arcpy.env.workspace=geodatabase+os.sep+dataset
        listaCapas = arcpy.ListFeatureClasses()

        for fc in listaCapas:
            describefc= arcpy.Describe(fc)
            if describefc.featureType != "Annotation":
                count = CountMoreZero(fc)
                print fc +"..."+ str(count)
                if count > 0:
                    ListaCapasGeo.append(geodatabase+os.sep+dataset+os.sep+fc)
    return ListaCapasGeo
def exportXLSX(Feat,ListFields):
    wb = Workbook()
    ws = wb.create_sheet(title=os.path.basename(Feat))
    print Feat
    with arcpy.da.SearchCursor(Feat, ListFields) as cursor:
        X=0
        for row in cursor:
            for col in range(0, len(ListFields)):
                try:
                    ws.cell(column=col, row=X, value=row[col])
                except:
                    pass
            X=X+1

    wb.save(filename = dest_filename)

def compararPuntos (buffer, capaEntrada, capaSalida, fields):
    featSalida= "in_memory\salida"
    join= arcpy.SpatialJoin_analysis(capaEntrada,capaSalida,featSalida,"JOIN_ONE_TO_ONE","KEEP_COMMON","","INTERSECT",buffer)
    exportXLSX(featSalida,fields)


codigos= {("2306","2101"):"Area de desecho",
("2325","2102"):"Cementerio",
("3686","2102"):"Cementerio",
("2325","2103"):"Central Telefonica",
("4101","2103"):"Central Telefonica",
("2325","2105"):"Faro",
("3413","2105"):"Faro",
("2304","2106"):"Instalacion de Energia",
("2309","2107"):"Zona Recreativa",
("2305","2108"):"Planta de Tratamiento",
("2307","2109"):"Zonas deportivas ",
("2325","2109"):"Zonas deportivas ",
("2325","2110"):"Estacion de transporte masivo",
("2325","2111"):"Estacion de Bombeo",
("4101","2111"):"Estacion de Bombeo",
("2325","2201"):"Construccion Habitacional",
("4101","2201"):"Construccion Habitacional",
("2325","2202"):"Construccion Salud",
("4119","2202"):"Construccion Salud",
("2325","2204"):"Construccion Educativa",
("4142","2204"):"Construccion Educativa",
("2325","2205"):"Construccion Anexa",
("4101","2205"):"Construccion Anexa",
("2325","2206"):"Construccion Comercio ",
("2320","2206"):"Construccion Comercio ",
("2325","2207"):"Construccion Seguridad",
("4166","2207"):"Construccion Seguridad",
("2325","2208"):"Construccion Religiosa",
("4129","2208"):"Construccion Religiosa",
("2325","2209"):"Edificios Administrativos",
("2320","2209"):"Edificios Administrativos",
("2325","2210"):"Entidad Financiera",
("2320","2210"):"Entidad Financiera",
("1104","2212"):"Hito Limite",
("2325","2214"):"Hotel",
("4131","2214"):"Hotel",
("2325","2215"):"Industrias",
("3680","2215"):"Industrias",
("2325","2216"):"Instalacion Minera",
("2302","2216"):"Instalacion Minera",
("2321","2219"):"Mina",
("2325","2221"):"Sitio de Interes",
("2320","2221"):"Sitio de Interes",
("2325","2223"):"Tanque",
("5610","2223"):"Tanque",
("2325","2224"):"Instalacion Petrolera",
("5620","2224"):"Instalacion Petrolera",
("2325","2226"):"Educacion no Formal",
("4101","2226"):"Educacion no Formal",
("4165","2227"):"Monumento",
("2325","2227"):"Monumento",
("3201","3201"):"Aeropuerto",
("3203","3202"):"Helipuerto",
("3202","3203"):"Pista de Aterrizaje",
("3301","3303"):"Puerto",
("2325","3304"):"Ferri",
("2320","3304"):"Ferri",
("3402","3402"):"Peaje",
("3401","3404"):"Terminal",
("3410","3406"):"Tunel",
("2325","4101"):"Area desertica",
("4165","4101"):"Area desertica",
("8132","4102"):"Glaciar",
("8100","4102"):"Glaciar",
("5106","4103"):"Humedal",
("5116","4104"):"Manglar",
("5129","4105"):"Morichal",
("8132","4106"):"Nevado",
("8100","4106"):"Nevado",
("8100","4107"):"PAramo",
("8135","4107"):"PAramo",
("8400","5101"):"Bahia",
("8403","5101"):"Bahia",
("8413","5102"):"Canal Marino",
("8400","5102"):"Canal Marino",
("5115","5103"):"Cienaga",
("8400","5104"):"Ensenada",
("8417","5104"):"Ensenada",
("8400","5105"):"Estrecho",
("8419","5105"):"Estrecho",
("8400","5106"):"Golfo",
("8422","5106"):"Golfo",
("8425","5107"):"Feb",
("8400","5107"):"Mar",
("8400","5108"):"Oceano",
("8427","5108"):"Oceano",
("5113","5301"):"Embalse",
("8400","5302"):"Estero",
("8418","5302"):"Estero",
("5128","5303"):"Pozo",
("5112","5304"):"Laguna",
("5105","5305"):"Manantial",
("5114","5306"):"Pantano",
("8400","5307"):"Rincon",
("8432","5307"):"Rincon",
("7101","7101"):"Republica",
("8317","7101"):"Republica",
("7101","7201"):"Area Metropolitana",
("8301","7201"):"Area Metropolitana",
("7101","7202"):"departamento",
("8309","7202"):"departamento",
("7101","7203"):"Provincia",
("8315","7203"):"Provincia",
("7101","7301"):"Corregimiento Departamental",
("8308","7301"):"Corregimiento Departamental",
("7101","7302"):"Distrito",
("8310","7302"):"Distrito",
("7101","7303"):"Municipio",
("8312","7303"):"Municipio",
("7101","7401"):"Cabecera Municipal",
("8304","7401"):"Cabecera Municipal",
("7101","7402"):"Capital",
("8324","7402"):"Capital",
("7101","7403"):"Centro Poblado",
("8326","7403"):"Centro Poblado",
("7101","7404"):"Corregimiento",
("8307","7404"):"Corregimiento",
("7101","7405"):"Localidad",
("8311","7405"):"Localidad",
("7101","7406"):"Region",
("8325","7406"):"Region",
("7101","7407"):"Zona Urbana",
("8326","7407"):"Zona Urbana",
("7101","7501"):"Barrio",
("8303","7501"):"Barrio",
("7101","7502"):"Inspeccion de Policia",
("8323","7502"):"Inspeccion de Policia",
("7101","7503"):"Sector",
("8327","7503"):"Sector",
("7101","7504"):"Vereda",
("8322","7504"):"Vereda",
("7101","7505"):"Urbanizacion",
("8350","7505"):"Urbanizacion",
("7101","7601"):"Caserio",
("8305","7601"):"Caserio",
("7101","7602"):"Comuna",
("8328","7602"):"Comuna",
("7101","7603"):"Sitios",
("8351","7603"):"Sitios",
("7101","7701"):"Rancheria",
("8326","7701"):"Rancheria",
("7101","7702"):"Reserva Indigena",
("8318","7702"):"Reserva Indigena",
("7101","7703"):"Resguardo Indigena",
("8320","7703"):"Resguardo Indigena",
("7101","7704"):"Tierras Colectivas de Comunidades Negras",
("8306","7704"):"Tierras Colectivas de Comunidades Negras",
("7101","7801"):"Area Natural unica",
("8302","7801"):"Area Natural unica",
("7101","7802"):"Parque Arqueologico",
("8313","7802"):"Parque Arqueologico",
("7101","7803"):"Parque Nacional Natural",
("8314","7803"):"Parque Nacional Natural",
("7101","7804"):"Reserva Forestal",
("8319","7804"):"Reserva Forestal",
("7101","7805"):"Reserva Natural",
("8319","7805"):"Reserva Natural",
("7101","7806"):"Santuario de Flora y Fauna",
("8321","7806"):"Santuario de Flora y Fauna",
("8100","9101"):"Altiplano",
("8101","9101"):"Altiplano",
("8100","9102"):"Bajo(TSI)",
("8104","9102"):"Bajo(TSI)",
("5201","9103"):"Banco Arena",
("8100","9104"):"Boqueron",
("8103","9104"):"Boqueron",
("8100","9105"):"Canon",
("8105","9105"):"Canon",
("8100","9106"):"Cerro",
("8106","9106"):"Cerro",
("8100","9107"):"Colina",
("8107","9107"):"Colina",
("8100","9108"):"Cordillera",
("8108","9108"):"Cordillera",
("8100","9109"):"CrAter",
("8109","9109"):"CrAter",
("8100","9110"):"Cuchilla",
("8100","9110"):"Cuchilla",
("8100","9111"):"Cueva",
("8111","9111"):"Cueva",
("8100","9112"):"Depresion",
("8112","9112"):"Depresion",
("8100","9113"):"Farallon",
("8114","9113"):"Farallon",
("5206","9114"):"Isla",
("8100","9115"):"Llanura",
("1816","9115"):"Llanura",
("8100","9116"):"Loma",
("8115","9116"):"Loma",
("8100","9117"):"Macizo",
("8117","9117"):"Macizo",
("8100","9118"):"Mesa",
("8118","9118"):"Mesa",
("8100","9119"):"Meseta",
("8919","9119"):"Meseta",
("8100","9120"):"Monte",
("8114","9120"):"Monte",
("8100","9121"):"Nudo",
("8121","9121"):"Nudo",
("8100","9122"):"Pena",
("8122","9122"):"Pena",
("8123","9123"):"Pico",
("8123","9123"):"Pico",
("8100","9124"):"Playon",
("8125","9124"):"Playon",
("8100","9125"):"Ramal",
("8126","9125"):"Ramal",
("8100","9126"):"Sabana",
("8127","9126"):"Sabana",
("8100","9127"):"Serrania",
("8128","9127"):"Serrania",
("8100","9128"):"Sierra",
("8129","9128"):"Sierra",
("8100","9129"):"Valle",
("8130","9129"):"Valle",
("8100","9130"):"VolcAn",
("8166","9130"):"VolcAn",
("8400","9201"):"Archipielago",
("8401","9201"):"Archipielago",
("8400","9202"):"Arrecife",
("8402","9202"):"Arrecife",
("8400","9203"):"Bajo En El Mar",
("8404","9203"):"Bajo En El Mar",
("8400","9204"):"Banco",
("8405","9204"):"Banco",
("8400","9205"):"Cabo",
("8411","9205"):"Cabo",
("8400","9206"):"Cayo",
("8414","9206"):"Cayo",
("5127","9207"):"Costa",
("8400","9208"):"Delta",
("8416","9208"):"Delta",
("5206","9209"):"Isla En El Mar",
("8400","9210"):"Islote",
("8423","9210"):"Islote",
("8400","9211"):"Istmo",
("8424","9211"):"Istmo",
("8400","9212"):"Peninsula",
("8428","9212"):"Peninsula",
("8400","9213"):"Penon",
("8429","9213"):"Penon",
("8400","9214"):"Playa",
("8430","9214"):"Playa",
("8400","9215"):"Punta",
("8431","9215"):"Punta"}

fields=["IDE","ENT_GEO","TOPONIMO","NOMBRE_GEOGRAFICO"]

toponimia= recorrerTopo(GeodatabaseTopo)
print toponimia
vectores =recorreGeo(GeodatabaseGeo)
print vectores
for fctop in toponimia:
    if "IB" == os.path.basename(fctop):
        for fcvec in vectores:
            if os.path.basename(fcvec)=="Construccion_P":
                compararPuntos("10 METERS",fctop,fcvec,fields)


