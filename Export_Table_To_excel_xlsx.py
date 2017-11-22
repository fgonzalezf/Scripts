#!/usr/bin/python
# -*- coding: utf-8 -*-
import arcpy, os, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
wb = Workbook()
dest_filename = r'C:\Users\APN\Documents\SGC\Muestras\Origen_Bogota\Libro1.xlsx'
wb.save(filename = dest_filename)
def campos(Feat):
    listadoCampos=[]
    fields=arcpy.ListFields(Feat)
    for field in fields:
        if field.editable==True and field.type!="Geometry":
            Temparr=[]
            Temparr.append(field.name)
            Temparr.append(field.type)
            Temparr.append(field.length)
            listadoCampos.append(Temparr)
    return listadoCampos


def exportXLSX(Feat,ListFields,wb):
    wb = load_workbook(dest_filename)
    ws = wb.create_sheet(title=os.path.basename(Feat))
    print Feat
    for col in range(1, len(ListFields)):
            dv=None
            if ListFields[col][1]=="String":
                dv = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=ListFields[col][2])
                dv.error ='Valor Invalido'
                dv.errorTitle = 'Error Texto'
                dv.prompt = "Texto menor a "+ str(ListFields[col][2])
                dv.promptTitle = ListFields[col][0]
                #print "string"
            elif ListFields[col][1]=="Double":
                dv = DataValidation(type="decimal")
                dv.error ='Valor Invalido'
                dv.errorTitle = 'Error Valor Decimal'
                dv.prompt = "Valor Decimal"
                dv.promptTitle = ListFields[col][0]

                #print "Double"
            elif ListFields[col][1]=="SmallInteger" or ListFields[col][1]=="Integer":
                dv = DataValidation(type="whole")
                dv.error ='Valor Invalido'
                dv.errorTitle = 'Error valor Entero'
                dv.prompt = "Valor Entero"
                dv.promptTitle = ListFields[col][0]

                #print "Entero"
            elif ListFields[col][1]=="Date":
                dv = DataValidation(type="date")
                dv.error ='Valor Invalido'
                dv.errorTitle = 'Error valor Fecha'
                dv.prompt = "Valor Fecha"
                dv.promptTitle = ListFields[col][0]

                #print "Date"
            ws.cell(column=col, row=1, value="{0}".format(ListFields[col][0].encode('utf-8')))
            #dv.ranges.append()
            for row in range(2,65563):
                #ws.cell(column=col, row=row).number_format=""
                dv.add(ws[ws.cell(column=col, row=row).coordinate])
            ws.add_data_validation(dv)
    wb.save(filename = dest_filename)






arcpy.env.workspace = r"C:\Users\APN\Documents\SGC\Muestras\Origen_Bogota\mg100K_Sin_Rel.gdb\Muestras"
featureclasses = arcpy.ListFeatureClasses()

for fc in featureclasses:
    #FeatEntrada =r"C:\Users\APN\Documents\SGC\Muestras\Origen_Bogota\mg100K_Sin_Rel.gdb\Muestras\Gravas"
    ListaCampos=campos(fc)
    exportXLSX(fc,ListaCampos,wb)

#dest_filename = r'C:\Users\APN\Documents\SGC\Muestras\Origen_Bogota\Libro1.xlsx'
#wb.save(filename = dest_filename)