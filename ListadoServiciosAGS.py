#!/usr/bin/python
#-*- coding: utf-8 -*-
#import restapi
from restapi import admin
import json
import arcpy, os, sys
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
wb = Workbook()
dest_filename = r'C:\temp\Libro1.xlsx'
wb.save(filename = dest_filename)

wb = load_workbook(dest_filename)
ws = wb.create_sheet(title="Servicios")
ListFields=["ServiceName","Folder","RecycleInterval","RecycleStartTime","Type","Status","Min Instances","Max Instances"]

url = 'srvags.sgc.gov.co/arcgis/rest/services' #server url
usr = r'ingeominas\fgonzalezf'
pw = 'Maidenfgf33'

# connect to ArcGIS Server instance
arcserver = admin.ArcServerAdmin(url, usr, pw)
row=1

def exportXLSX(ListFields,wb, row):
    for service in arcserver.iter_services():
        print service
        #data = json.load(service)
        for col in range(1, len(ListFields)+1):
            if row==1:
                if ListFields[col-1]=="ServiceName":
                    ws.cell(column=col, row=row, value="{0}".format("ServiceName"))
                elif ListFields[col-1]=="Folder":
                    ws.cell(column=col, row=row, value="{0}".format("Folder"))
                elif ListFields[col - 1] == "RecycleInterval":
                    ws.cell(column=col, row=row, value="{0}".format("RecycleInterval"))
                elif ListFields[col-1]=="RecycleStartTime":
                    ws.cell(column=col, row=row, value="{0}".format("RecycleStartTime"))
                elif ListFields[col-1] == "Type":
                    ws.cell(column=col, row=row, value="{0}".format("Type"))
                elif ListFields[col-1] == "Status":
                    ws.cell(column=col, row=row, value="{0}".format("Status"))
                elif ListFields[col-1] == "Min Instances":
                    ws.cell(column=col, row=row, value="{0}".format("Min Instances"))
                elif ListFields[col-1] == "Max Instances":
                    ws.cell(column=col, row=row, value="{0}".format("Max Instances"))
            else:
                if ListFields[col-1]=="ServiceName":
                    ws.cell(column=col, row=row, value="{0}".format(service["serviceName"].encode('utf-8')))
                elif ListFields[col - 1] == "Folder":
                    try:
                        ws.cell(column=col, row=row, value="{0}".format(service["properties"]["filePath"].replace("E:\\arcgisserver\\directories\\arcgissystem\\arcgisinput\\","").split("\\")[0]))
                    except Exception as e:
                        print e.message
                elif ListFields[col-1]=="RecycleInterval":
                    ws.cell(column=col, row=row, value="{0}".format(service["recycleInterval"]))
                elif ListFields[col-1]=="RecycleStartTime":
                    ws.cell(column=col, row=row, value="{0}".format(service["recycleStartTime"]))
                elif ListFields[col-1] == "Type":
                    ws.cell(column=col, row=row, value="{0}".format(service["type"].encode('utf-8')))
                elif ListFields[col-1] == "Status":
                    ws.cell(column=col, row=row, value="{0}".format(service["configuredState"].encode('utf-8')))
                elif ListFields[col-1] == "Min Instances":
                    ws.cell(column=col, row=row, value="{0}".format(service["minInstancesPerNode"]))
                elif ListFields[col-1] == "Max Instances":
                    ws.cell(column=col, row=row, value="{0}".format(service["maxInstancesPerNode"]))
        row = row + 1
    wb.save(filename = dest_filename)

exportXLSX(ListFields,wb, row)



